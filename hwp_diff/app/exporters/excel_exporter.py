"""Excel exporter for comparison results — uses QAR template (template2.xlsx)."""
import re
from pathlib import Path
from typing import List, Optional

from app.models.change_record import ChangeRecord, CompareResult, ChangeType, Importance
from app.utils.logger import get_logger

logger = get_logger("exporters.excel")

DEFAULT_TEMPLATE = r"F:\template2.xlsx"
QAR_SHEET = "QAR"
DATA_START_ROW = 7  # template의 헤더는 row 1~6, 데이터는 row 7부터

# 컬럼 인덱스 (1-based)
COL_SEQ   = 1   # A: 순위
COL_PROD  = 2   # B: 재고번호(품명)
COL_SPEC  = 3   # C: 규격서번호/도번
COL_COORD = 4   # D: 규격항목명/좌표  →  "X매중 Y매"
COL_OLD   = 5   # E: ~을(변경전)
COL_NEW   = 6   # F: ~으로(변경후)
COL_SUM1  = 7   # G: 요약(제안필요성)  →  "○ ... 내용명확화\n-, 영향을 미치는사항"
COL_SUM2  = 8   # H: 요약(영향)        →  "영향없음"
COL_CODE  = 9   # I: 변경사유코드      →  "T6"


class ExcelExporter:
    """Exports CompareResult to an Excel workbook using QAR template."""

    def export(
        self,
        result: CompareResult,
        output_path: str,
        old_file_name: str = "기준문서",
        new_file_name: str = "비교문서",
        product_name: str = "",
        spec_number: str = "",
        template_path: str = "",
        total_pages: int = 0,
    ) -> bool:
        """
        Export comparison result to Excel using QAR template.

        Args:
            result: CompareResult from DiffEngine.compare()
            output_path: Path to output .xlsx file
            old_file_name: Display name for the original document
            new_file_name: Display name for the new document
            product_name: 품명 (from UI)
            spec_number: 규격서번호 또는 도번 (from UI)
            template_path: QAR template .xlsx path (default: F:\\template2.xlsx)
            total_pages: 총 페이지 수 (for "X매중 Y매" format)

        Returns:
            True on success, False on failure
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl is not installed. Cannot export Excel.")
            return False

        tpl = template_path or DEFAULT_TEMPLATE

        try:
            # ── 템플릿 로드 ───────────────────────────────────────────────
            if Path(tpl).exists():
                wb = openpyxl.load_workbook(tpl)
                logger.info("Loaded QAR template: %s", tpl)
            else:
                logger.warning("Template not found (%s) — creating blank workbook", tpl)
                wb = self._create_blank_wb()

            # QAR 시트 선택
            if QAR_SHEET in wb.sheetnames:
                ws = wb[QAR_SHEET]
            else:
                ws = wb.active
                ws.title = QAR_SHEET

            # ── 총 페이지 수 계산 ─────────────────────────────────────────
            if total_pages <= 0:
                pages = [_extract_page_num(c.page_hint) for c in result.changes]
                pages = [p for p in pages if p > 0]
                total_pages = max(pages) if pages else 1

            # ── 기존 데이터 행 초기화 (양식 재사용 시 잔여 데이터 제거) ──
            last_row = ws.max_row
            if last_row >= DATA_START_ROW:
                for r in range(DATA_START_ROW, last_row + 1):
                    for c in range(1, 10):
                        ws.cell(row=r, column=c).value = None

            # ── 변경사항 행 기록 ──────────────────────────────────────────
            for seq, change in enumerate(result.changes, 1):
                row_idx = DATA_START_ROW + seq - 1
                ct = change.change_type

                # A: 순위
                ws.cell(row=row_idx, column=COL_SEQ).value = seq

                # B: 품명
                ws.cell(row=row_idx, column=COL_PROD).value = product_name or ""

                # C: 규격서번호/도번
                ws.cell(row=row_idx, column=COL_SPEC).value = spec_number or ""

                # D: 좌표 → "X매중 Y매"
                page_num = _extract_page_num(change.page_hint)
                if page_num <= 0:
                    page_num = 1
                ws.cell(row=row_idx, column=COL_COORD).value = f"{total_pages}매중 {page_num}매"

                # E: 변경 전 내용
                if ct == ChangeType.ADD:
                    old_val = ""
                else:
                    old_val = change.old_content or ""
                ws.cell(row=row_idx, column=COL_OLD).value = old_val

                # F: 변경 후 내용 (추가/삭제 마커 포함)
                if ct == ChangeType.ADD:
                    new_val = f"추가  {change.new_content}" if change.new_content else "추가"
                elif ct == ChangeType.DELETE:
                    new_val = "삭제"
                else:
                    new_val = change.new_content or ""
                ws.cell(row=row_idx, column=COL_NEW).value = new_val

                # G: 요약(제안필요성) — "○ [내용] 내용명확화\n-, 영향을 미치는사항"
                detail = change.summary or change.location_info or ""
                ws.cell(row=row_idx, column=COL_SUM1).value = (
                    f"○ {detail} 내용명확화\n-, 영향을 미치는사항"
                )

                # H: 요약(영향)
                ws.cell(row=row_idx, column=COL_SUM2).value = "영향없음"

                # I: 변경사유코드
                ws.cell(row=row_idx, column=COL_CODE).value = "T6"

                # 스타일 적용
                self._style_row(ws, row_idx)

            # ── 저장 ─────────────────────────────────────────────────────
            output_path_obj = Path(output_path)
            output_path_obj.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(output_path_obj))
            logger.info("Excel exported: %s", output_path)
            return True

        except Exception as e:
            logger.exception("Excel export failed: %s", e)
            return False

    # ------------------------------------------------------------------
    # Style helpers
    # ------------------------------------------------------------------

    def _style_row(self, ws, row_idx: int) -> None:
        """Apply border / alignment to a single data row."""
        from openpyxl.styles import Font, Alignment, Border, Side

        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if col_idx in (COL_OLD, COL_NEW, COL_SUM1):
                # 내용 컬럼: 왼쪽 정렬 + 줄바꿈
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cell.font = Font(size=9)
            elif col_idx == COL_SEQ:
                # 순위: 가운데
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)
            else:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.font = Font(size=9)

        ws.row_dimensions[row_idx].height = 55

    # ------------------------------------------------------------------
    # Fallback blank workbook (template 없을 때)
    # ------------------------------------------------------------------

    def _create_blank_wb(self):
        """Create a minimal QAR workbook when template file is unavailable."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = QAR_SHEET

        # 타이틀 (row 1)
        ws.merge_cells("A1:I1")
        ws["A1"] = "세부항목내역서 (QAR)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 헤더 (row 6 — DATA_START_ROW - 1)
        headers = [
            "순위", "품명", "규격서번호/도번",
            "좌표", "변경전(~을)", "변경후(~으로)",
            "요약(제안필요성)", "요약(영향)", "변경사유코드",
        ]
        hdr_row = DATA_START_ROW - 1
        hdr_fill = PatternFill("solid", fgColor="D9D9D9")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=hdr_row, column=col_idx, value=hdr)
            cell.font = Font(bold=True, size=9)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[hdr_row].height = 25

        # 컬럼 너비
        col_widths = {1: 6, 2: 18, 3: 18, 4: 14, 5: 38, 6: 38, 7: 38, 8: 14, 9: 14}
        for ci, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w

        return wb


# ------------------------------------------------------------------
# Module-level helpers
# ------------------------------------------------------------------

def _extract_page_num(page_hint) -> int:
    """
    page_hint 값에서 페이지 번호(정수)를 추출한다.

    지원 형식:
      - int / float  → 그대로 int 변환
      - "3페이지 (추정)"  → 3
      - "페이지 미상"      → 0
      - "5"               → 5
    """
    if page_hint is None:
        return 0
    if isinstance(page_hint, (int, float)):
        return int(page_hint)
    s = str(page_hint).strip()
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0
