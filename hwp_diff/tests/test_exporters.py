"""Tests for the Excel exporter."""
import os
import sys
import tempfile
import unittest
import uuid
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from app.models.change_record import (
    ChangeRecord, CompareResult, ChangeType, ObjectType, Importance
)
from app.exporters.excel_exporter import ExcelExporter


def _make_change(
    change_type: ChangeType,
    object_type: ObjectType = ObjectType.PARAGRAPH,
    importance: Importance = Importance.MEDIUM,
    old_content: str = "기존 내용",
    new_content: str = "새 내용",
) -> ChangeRecord:
    return ChangeRecord(
        change_id=f"CHG-{uuid.uuid4().hex[:6]}",
        change_type=change_type,
        object_type=object_type,
        location_info="3장 2절 문단 5",
        page_hint="3페이지 (추정)",
        old_content=old_content,
        new_content=new_content,
        summary=f"{change_type.value} 변경 요약",
        similarity=0.75,
        importance=importance,
        review_needed=(importance == Importance.HIGH),
    )


def _make_result() -> CompareResult:
    result = CompareResult()
    result.changes = [
        _make_change(ChangeType.MODIFY, importance=Importance.HIGH,
                     old_content="허용오차는 ±0.5V 이내이어야 한다.",
                     new_content="허용오차는 ±0.3V 이내이어야 한다."),
        _make_change(ChangeType.ADD, object_type=ObjectType.PARAGRAPH,
                     old_content="",
                     new_content="새로 추가된 문단입니다."),
        _make_change(ChangeType.DELETE, object_type=ObjectType.PARAGRAPH,
                     old_content="삭제된 문단입니다.",
                     new_content=""),
        _make_change(ChangeType.MOVE, importance=Importance.LOW,
                     old_content="이동된 내용",
                     new_content="이동된 내용"),
        _make_change(ChangeType.MODIFY, object_type=ObjectType.HEADING,
                     importance=Importance.HIGH,
                     old_content="1. 개요",
                     new_content="1. 개요 및 목적"),
        _make_change(ChangeType.MODIFY, object_type=ObjectType.CELL,
                     importance=Importance.HIGH,
                     old_content="25",
                     new_content="30"),
    ]
    result.total_old_blocks = 20
    result.total_new_blocks = 21
    result.added_count = 1
    result.deleted_count = 1
    result.modified_count = 3
    result.moved_count = 1
    result.format_count = 0
    result.table_change_count = 1
    result.heading_change_count = 1
    result.overall_similarity = 0.82
    result.revision_rate = 0.18
    return result


class TestExcelExporter(unittest.TestCase):
    """Tests for ExcelExporter."""

    def setUp(self):
        self.exporter = ExcelExporter()

    def test_export_creates_file(self):
        """Export creates an xlsx file."""
        result = _make_result()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            success = self.exporter.export(result, path)
            self.assertTrue(success, "Export should return True on success")
            self.assertTrue(os.path.exists(path), "Output file should exist")
            self.assertGreater(os.path.getsize(path), 0, "Output file should not be empty")
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_has_correct_sheets(self):
        """Exported workbook has the expected sheet names."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            sheet_names = wb.sheetnames
            self.assertIn("전체 변경사항", sheet_names)
            self.assertIn("표 변경사항", sheet_names)
            self.assertIn("제목_목차 변경", sheet_names)
            self.assertIn("요약 통계", sheet_names)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_all_changes_sheet_row_count(self):
        """전체 변경사항 sheet has correct number of data rows."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["전체 변경사항"]
            # Count non-None rows that contain a sequence number in column 1
            data_rows = sum(
                1 for row in ws.iter_rows(min_row=3, values_only=True)
                if row[0] is not None and isinstance(row[0], int)
            )
            self.assertEqual(data_rows, len(result.changes),
                             f"Expected {len(result.changes)} data rows, got {data_rows}")
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_table_changes_filtered(self):
        """표 변경사항 sheet contains only table/cell changes."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        table_changes = [
            c for c in result.changes
            if c.object_type.value in ("표", "셀")
        ]
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["표 변경사항"]
            data_rows = sum(
                1 for row in ws.iter_rows(min_row=3, values_only=True)
                if row[0] is not None and isinstance(row[0], int)
            )
            self.assertEqual(data_rows, len(table_changes))
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_heading_changes_filtered(self):
        """제목_목차 변경 sheet contains only heading changes."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        heading_changes = [c for c in result.changes if c.object_type.value == "제목"]
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["제목_목차 변경"]
            data_rows = sum(
                1 for row in ws.iter_rows(min_row=3, values_only=True)
                if row[0] is not None and isinstance(row[0], int)
            )
            self.assertEqual(data_rows, len(heading_changes))
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_empty_result(self):
        """Handles empty result gracefully."""
        result = CompareResult()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            success = self.exporter.export(result, path)
            self.assertTrue(success)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_korean_content(self):
        """Verify Korean text is preserved in export."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = CompareResult()
        result.changes = [
            _make_change(
                ChangeType.MODIFY,
                old_content="허용오차는 ±0.5V 이내이어야 한다.",
                new_content="허용오차는 ±0.3V 이내이어야 한다.",
            )
        ]
        result.modified_count = 1

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["전체 변경사항"]
            # Data starts at row 3 (row 1=title, row 2=header)
            # Column 7 = old content
            found_korean = False
            for row in ws.iter_rows(min_row=3, values_only=True):
                for cell_val in row:
                    if cell_val and "허용오차" in str(cell_val):
                        found_korean = True
                        break
            self.assertTrue(found_korean, "Korean content should be preserved in Excel")
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_summary_stats_present(self):
        """요약 통계 sheet contains key statistics."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self.exporter.export(result, path)
            wb = openpyxl.load_workbook(path)
            ws = wb["요약 통계"]

            # Collect all cell values
            all_values = set()
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    if val is not None:
                        all_values.add(str(val))

            self.assertIn("총 변경", all_values)
            self.assertIn("추가", all_values)
            self.assertIn("삭제", all_values)
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_export_no_output_path_fails(self):
        """Returns False when no output path provided."""
        result = _make_result()
        success = self.exporter.export(result, "")
        self.assertFalse(success)

    def test_export_creates_parent_dirs(self):
        """Creates parent directories if they don't exist."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        result = _make_result()
        tmp_dir = tempfile.mkdtemp()
        nested_path = os.path.join(tmp_dir, "subdir", "output.xlsx")
        try:
            success = self.exporter.export(result, nested_path)
            self.assertTrue(success)
            self.assertTrue(os.path.exists(nested_path))
        finally:
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
